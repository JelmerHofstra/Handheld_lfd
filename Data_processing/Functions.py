from rosbag2_py import SequentialReader, StorageOptions, ConverterOptions
from rosidl_runtime_py.utilities import get_message
from sensor_msgs.msg import Imu
from rclpy.serialization import deserialize_message
from cv_bridge import CvBridge
import sensor_msgs_py.point_cloud2 as pc2
from sensor_msgs.msg import PointCloud2
import cv2
import struct
import os
import time
import json
from tqdm import tqdm
import openpyxl
from openpyxl import Workbook
import numpy as np
import pandas as pd
from scipy.spatial.transform import Rotation as R
from openpyxl import load_workbook
import matplotlib.pyplot as plt


def Makereader(bag_path) :
    reader = SequentialReader()
    storage_options = StorageOptions(uri=bag_path, storage_id='sqlite3')
    converter_options = ConverterOptions(input_serialization_format='cdr', output_serialization_format='cdr')
    reader.open(storage_options, converter_options)
    topic_types = reader.get_all_topics_and_types()
    type_map = {t.name: t.type for t in topic_types}
    return topic_types , reader , type_map


def Analysebag(output_dir,bag_path, topic_pointcloud):
    topic_types , reader,_ = Makereader(bag_path)
    for topic in topic_types :
        print(topic.name)
    for topic in topic_types: 
        if topic.name == topic_pointcloud:
            Pointcloud_avaible = True
            break
        else :
            Pointcloud_avaible = False

    imu_type = Imu  # We know the type already
    while reader.has_next():
        topic, data, _ = reader.read_next()

        if topic == '/cam_ext/zed/imu/data' :
            msg = imu_type()
            msg = deserialize_message(data, Imu)            
            cam_ori = [ msg.orientation.x,  msg.orientation.y, msg.orientation.z,  msg.orientation.w]
            break

    with open(f"{output_dir}/configs.txt", "w") as file:
        file.write(f"Pointcloud_avaible: {Pointcloud_avaible}\n")
        file.write(f"Camera Orientation: {cam_ori}\n")

    
    return Pointcloud_avaible ,cam_ori







def deserialize_msg(topic, raw, type_map):
    msg_type = get_message(type_map[topic])
    msg = deserialize_message(raw, msg_type)
    return msg


def image_msg_to_cv2(msg,bridge):
    return bridge.imgmsg_to_cv2(msg, desired_encoding='bgr8')


def detect_aruco_markers(cv_image):
    aruco_dict = cv2.aruco.Dictionary_get(cv2.aruco.DICT_4X4_50)
    parameters = cv2.aruco.DetectorParameters_create()
    corners, ids, _ = cv2.aruco.detectMarkers(cv_image, aruco_dict, parameters=parameters)
    return corners, ids 



def Makeanalyse_files(bag_path, topic_image,name_bag_file,topic_pointcloud,imu_topic,wrench_topic, image_int_topic):
    bridge = CvBridge()
    topic_types, reader, type_map = Makereader(bag_path)
    

    output_dir = f'./Data_processing/Recordings/{name_bag_file}'
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir,'video.mp4')
    output_path_int= os.path.join(output_dir,'video_intern.mp4')
    Pointcloud_avaible ,_ = Analysebag(output_dir,bag_path, topic_pointcloud)

    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    fps = 30

    video_writer = None
    video_writer_int = None
    framenumbers_skip = -1
    detected = 0
    frame_number = 0
    aruco_detections = []
    imu_data = []
    wrench_data = []
    camera_int_time =[]
    t0 = None

    while reader.has_next():
        topic, data, t = reader.read_next()

        if t0 is None:
            t0 = t


        if topic == topic_image:
            
           
            img_msg = deserialize_msg(topic, data, type_map)
            cv_image = image_msg_to_cv2(img_msg, bridge)
            framenumbers_skip +=1 
            if video_writer is None:
                height, width = cv_image.shape[:2]
                frame_size = (width, height)
                video_writer = cv2.VideoWriter(output_path, fourcc, fps, frame_size)
                if not video_writer.isOpened():
                    print("[ERROR] Could not open video writer.")
                    return

            gray = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
            corners, ids = detect_aruco_markers(gray)
            
            if ids is not None and len(ids) > 0:
                

                areas = []
                for corner in corners:
                    polygon = corner.reshape((4, 2))
                    area = cv2.contourArea(polygon)
                    areas.append(area)
                max_index = areas.index(max(areas))
                largest_id = int(ids.flatten()[max_index])
                largest_corners = corners[max_index].reshape(4, 2)

                if largest_id > 5 : 
                    continue

                detected += 1
                frame_number += 1

                u , v = np.mean(largest_corners, axis=0) 
                centrepoint = [u/width , v/height]
                elapsed_time = (t - t0) / 1e9

                aruco_detections.append({
                    "frame_number": frame_number,
                    "corners": largest_corners,
                    "Centrepoint" : centrepoint,
                    "id": largest_id,
                    "timestamp": elapsed_time,
                    "original_frame": framenumbers_skip
                })

                video_writer.write(cv_image)
        elif topic == imu_topic:
            imu_msg = deserialize_msg(topic, data, type_map)
            imu_data.append({
                "timestamp": (t - t0)/1e9,
                "orientation_x": imu_msg.orientation.x,
                "orientation_y": imu_msg.orientation.y,
                "orientation_z": imu_msg.orientation.z,
                "orientation_w": imu_msg.orientation.w,
                })
            
        elif topic == wrench_topic:
            wrench_msg = deserialize_msg(topic, data, type_map)
            # Process wrench data if needed
            wrench_data.append({
                "timestamp": (t - t0)/1e9,
                "force_x": wrench_msg.wrench.force.x,
                "force_y": wrench_msg.wrench.force.y,
                "force_z": wrench_msg.wrench.force.z,
                "torque_x": wrench_msg.wrench.torque.x,
                "torque_y": wrench_msg.wrench.torque.y,
                "torque_z": wrench_msg.wrench.torque.z,
            })

        elif topic == image_int_topic :
            img_msg = deserialize_msg(topic, data, type_map)
            cv_image = image_msg_to_cv2(img_msg, bridge)
            if video_writer_int is None:
                height_int, width_int = cv_image.shape[:2]
                frame_size_int = (width_int, height_int)
                fps_int = 15
                video_writer_int = cv2.VideoWriter(output_path_int, fourcc, fps, frame_size_int)
                if not video_writer_int.isOpened():
                    print("[ERROR] Could not open video writer.")
                    return

            video_writer_int.write(cv_image)
        
            camera_int_time.append({
                "timestamp": (t - t0)/1e9,
            })


        

           
    if video_writer:
        video_writer.release()
        print(f"✅ Video saved to {output_path}")
    else:
        print("⚠️ No video was created. Check if the topic contained images.")

    if video_writer_int:
        video_writer_int.release()
        print(f"✅ Internal video saved to {output_path_int}")

    save_detections_to_excel(aruco_detections, output_dir + '/Data.xlsx')
    save_imu_data_to_excel(imu_data, output_dir + '/imu_Data.xlsx')
    save_force_torque_data_to_excel(wrench_data, output_dir + '/wrench_Data.xlsx')
    save_camera_int_data_to_excel(camera_int_time, output_dir + '/camera_int_time.xlsx')
    return Pointcloud_avaible 


def save_detections_to_excel(aruco_detections, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["frame_number", "timestamp", "id", "corners" , 'Centrepoint'])

    for det in aruco_detections:
        frame_number = det.get("frame_number", None)
        timestamp = det["timestamp"]
        aruco_id = det["id"]
        corners = det["corners"]
        corners_str = json.dumps(corners.tolist())
        centrepoint = det["Centrepoint"]
        centrepoint_str = json.dumps(centrepoint)
        original_number = det["original_frame"]
        ws.append([frame_number, timestamp, aruco_id, corners_str, centrepoint_str, original_number])

    wb.save(excel_path)
    print(f"📄 ArUco detection log saved to {excel_path}")

def save_imu_data_to_excel(imu_data, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["timestamp", "orientation_x", "orientation_y", "orientation_z", "orientation_w"])

    for imu in imu_data:
        timestamp = imu["timestamp"]
        orientation_x = imu["orientation_x"]
        orientation_y = imu["orientation_y"]
        orientation_z = imu["orientation_z"]
        orientation_w = imu["orientation_w"]
        ws.append([timestamp, orientation_x, orientation_y, orientation_z, orientation_w])

    wb.save(excel_path)
    print(f"📄 IMU data saved to {excel_path}")

def save_force_torque_data_to_excel(wrench_data, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["timestamp", "force_x", "force_y", "force_z", "torque_x", "torque_y", "torque_z"])

    for wrench in wrench_data:
        timestamp = wrench["timestamp"]
        force_x = wrench["force_x"]
        force_y = wrench["force_y"]
        force_z = wrench["force_z"]
        torque_x = wrench["torque_x"]
        torque_y = wrench["torque_y"]
        torque_z = wrench["torque_z"]
        ws.append([timestamp, force_x, force_y, force_z, torque_x, torque_y, torque_z])

    wb.save(excel_path)
    print(f"📄 Force and Torque data saved to {excel_path}")

    
def save_camera_int_data_to_excel(camera_int_time, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["timestamp"])

    for cam_time in camera_int_time:
        timestamp = cam_time["timestamp"]
        ws.append([timestamp])

    wb.save(excel_path)
    print(f"📄 Camera internal time data saved to {excel_path}")




def  Open_excel(name) :
    # Load workbook and active worksheet
    wb = load_workbook(f'./Data_processing/Recordings/{name}/Data.xlsx')
    ws = wb.active

    # Skip header row and extract timestamp and id
    framenumbers = []
    timestamps = []
    ids = []
    corners_list = []  # To store corner coordinates as NumPy arrays
    centrepoints_list = []   # To store centrepoint coordinates as NumPy arrays
    original_frame_list = []

    # Read rows
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        
        framenumber = int(row[0])
        timestamp = float(row[1])  # first column: timestamp
        aruco_id = int(row[2])      # second column: id
        corners_str = row[3]        # third column: corners as JSON string
        centrepoint_str = row[4]    # fourth column: centrepoint as JSON string
        orignial_frame = row[5] 

        # Convert corners_str back to numpy array
        try:
            corners = np.array(json.loads(corners_str), dtype=np.float32)
            centrepoint = np.array(json.loads(centrepoint_str), dtype=np.float32)
        except Exception as e:
            print(f"[ERROR] Could not parse corners at row {i+1}: {e}")
            corners = None

        # Append to lists
        framenumbers.append(framenumber)
        timestamps.append(timestamp)
        ids.append(aruco_id)
        corners_list.append(corners)
        centrepoints_list.append(centrepoint)
        original_frame_list.append(orignial_frame)

    return framenumbers,timestamps, ids ,corners_list, centrepoints_list, original_frame_list




def Getcentreposition(RotPlane,XYZ_coord : np.array ,distance2markers : int , XYZ_prev) :
    vx = RotPlane[0]-RotPlane[1]   #Vector x (Upper right to upperleft)
    vy = RotPlane[1]-RotPlane[2]   #Vector y  (Upper left to down left)
    perp = np.cross(vx,vy)
    normPerp = -perp/np.linalg.norm(perp)
    closest = float('inf')

    if XYZ_prev is None :
        
        if normPerp[2] > 0 :
            Centreposition =XYZ_coord + 0.5*distance2markers*normPerp
            Centreposition_alt = [ 0,0,0]
        else :
            Centreposition= XYZ_coord - 0.5*distance2markers*normPerp
            Centreposition_alt = [ 0,0,0]
    else : 
        for i in [-1, 1 ] :
            position = XYZ_coord + i*0.5*distance2markers*normPerp
            position_alt = XYZ_coord -i*0.5*distance2markers*normPerp
            distance = np.linalg.norm(position - XYZ_prev) 
            if distance < closest :
                closest = distance 
                Centreposition = position
                Centreposition_alt = position_alt


    
    return Centreposition, Centreposition_alt


def GetFacePoints(Rotationmatrix : np.array ,XYZ_coord : np.array ,SizeCube : int) :
    NormalPlane = np.array([[-.5*SizeCube , .5*SizeCube, 0],[.5*SizeCube , .5*SizeCube, 0 ], [.5*SizeCube , -.5*SizeCube, 0 ],[-.5*SizeCube , -.5*SizeCube, 0 ]]).T #Coordinate UpperLeft, UpperRight, BottomRight, BottomLeft   
    rot_plane = Rotationmatrix @ NormalPlane + np.ones((3, 4)) * XYZ_coord.reshape(-1, 1) #Rotates the plane and add origin.
    return rot_plane.T 


def GetOrientationCube(ArucoId : int ,RotPlane):
    if ArucoId == 0:
        Yvec = RotPlane[1]-RotPlane[0]
        Zvec = RotPlane[3]-RotPlane[0]
        Xvec =  np.cross(Yvec, Zvec)
    
    
    elif ArucoId == 1:
        Xvec = RotPlane[0]-RotPlane[1]
        Zvec = RotPlane[3]-RotPlane[0]
        Yvec = np.cross(Zvec,Xvec)
        
    
    elif ArucoId == 2:
        Yvec = RotPlane[1]-RotPlane[0]
        Zvec = RotPlane[3]-RotPlane[0]
        Xvec = np.cross(Yvec,Zvec)
    
        
    elif ArucoId == 3:
        Xvec = RotPlane[1]-RotPlane[0]
        Zvec = RotPlane[3]-RotPlane[0]
        Yvec = np.cross(Zvec,Xvec)


    elif ArucoId == 4:
        Yvec = RotPlane[0]-RotPlane[1]
        Xvec = RotPlane[3]-RotPlane[0]
        Zvec = np.cross(Xvec,Yvec)
        
    else:
        raise ValueError(f"Unsupported ArucoId: {ArucoId}")

    
    # Normalize vectors
    Xnorm = Xvec / np.linalg.norm(Xvec)
    Ynorm = Yvec / np.linalg.norm(Yvec)
    Znorm = Zvec / np.linalg.norm(Zvec)
    
    
    # Create the rotation matrix
    R_matrix = np.column_stack((Xnorm.T, Ynorm.T, Znorm.T))
    # Convert the rotation matrix to quaternion
    
    return  R_matrix



def GetRightorientation(marker_id, rvecs, tvecs, R_prev):
    best = float('-inf')
    skip = False
    for i in range(len(rvecs)):
        Rnow ,_ = cv2.Rodrigues(rvecs[i])
        if marker_id == 0 : 
            X_prev = R_prev[:,0]
            X_now = Rnow[:,0]
            orient = np.dot(X_prev,X_now) 
            if  orient > best:
                best = orient
                index = i 
                skip = False


        elif marker_id == 1 :
            Y_prev = R_prev[:,1]
            Y_now = Rnow[:,1]
            orient = np.dot(Y_prev,Y_now)
            if  orient > best:
                best = orient
                index = i 
                skip = False
                



        elif marker_id == 2 :
            X_prev = R_prev[:,0]
            X_now = Rnow[:,0]
            orient = np.dot(X_prev,X_now) 
            if orient > best:
               best = orient
               index = i 
               skip = False

            
        elif marker_id == 3 :
            Y_prev = R_prev[:,1]
            Y_now = Rnow[:,1]
            orient =np.dot(Y_prev,Y_now)
            if orient > best:
               best = orient
               index = i 
               skip = False

        elif marker_id == 4 :
            Z_prev = R_prev[:,2]
            Z_now = Rnow[:,2]
            orient = np.dot(Z_prev,Z_now)
            if orient > best:
               best = orient
               index = i 
               skip = False
        else :
            skip = True
            print('Error, no such face')
            

    if skip :
        Rnow = 0
        R_alt = 0
        index = 0

    else :
        Rnow ,_ = cv2.Rodrigues(rvecs[index])
        if index == 1:
            R_alt,_ = cv2.Rodrigues(rvecs[0])
        else :
            R_alt,_ = cv2.Rodrigues(rvecs[1])
  

    if isinstance(tvecs, np.ndarray):
        tvec = tvecs
    else :
        tvec = tvecs[index]
    


    return Rnow, R_alt,  skip , tvec


def GetCentreCube(rotm_aruco_face, Coordinate, MarkerSize,distance2markers,XYZ_prev) :
    Plane = GetFacePoints(rotm_aruco_face, Coordinate, MarkerSize)
    CentreCube,CentreCube_alt = Getcentreposition(Plane,Coordinate,distance2markers,XYZ_prev)
    return CentreCube ,CentreCube_alt, Plane






class Pose_PositionData:

    def __init__(self, CameraMatrix,DistCoeffs,object_points, MarkerSize,distance2markers, name,Pointcloud_avaible, bagpath, topic_pointcloud):
        self.CameraMatrix = CameraMatrix
        self.DistCoeffs = DistCoeffs
        self.object_points = object_points
        self.MarkerSize = MarkerSize
        self.distance2markers = distance2markers
        self.name = name
        self.skips = []
        self.exist = os.path.exists(f'./Data_processing/savings/{self.name}/Data_processed.xlsx') 
        self.ReadConfig()
        self.Pointcloud_avaible = Pointcloud_avaible
        self.bagpath = bagpath
        self.topic_pointcloud = topic_pointcloud


        if self.exist:
            answer = input(' Do you want to continue with your last data_processing session? [Y/n]')
            if answer.upper() == "Y" or answer.lower() == 'y' :
                self.LoadData()
                
            elif answer.upper() == "N" or answer.lower() == 'n' : 
                self.Calculate_pos_and_orient_first_go()
            
            else :
                print('Error')
        else :
            self.Calculate_pos_and_orient_first_go()





        


    def Calculate_pos_and_orient_first_go(self) : 
        _,timestamps, ids , corners, centrepoints, original_frames = Open_excel(self.name)

        if self.Pointcloud_avaible:
            self.GetDepthmapdepthdata(centrepoints,original_frames)
        
        frame_number = 0
        Data = []


        cap = cv2.VideoCapture(f'./Data_processing/Recordings/{self.name}/video.mp4')
        self.total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        while cap.isOpened() and frame_number < self.total_frames:
            ret, frame = cap.read()
            if not ret:
                break

            image_points = corners[frame_number]
            marker_id = ids[frame_number]
        

            if frame_number == 0:
                _, rvecs, tvecs, _ = cv2.solvePnPGeneric(self.object_points, image_points, self.CameraMatrix, self.DistCoeffs,flags=cv2.SOLVEPNP_IPPE_SQUARE)
                first_image = frame.copy()
                cv2.drawFrameAxes(first_image, self.CameraMatrix, self.DistCoeffs, rvecs[0], tvecs[0], 0.1)
                cv2.imshow(' Is this R handed frame? [y/n]',first_image)
                while True:
                    key = cv2.waitKey(0) & 0xFF  # Wait indefinitely for a key press

                    if key == ord('y'):
                        rvec = rvecs[0]
                        tvec = tvecs[0]
                        break
                    elif key == ord('n'):
                        rvec = rvecs[1]
                        tvec = tvecs[1]
                        break
                    elif key == 27:  # ESC to quit without response
                        print("User exited.")
                        break

                cv2.destroyAllWindows()

                Rotm_face, _ = cv2.Rodrigues(rvec)
                x, y, z = tvec.flatten() 
                CentreCube,CentreCube_alt, RotPlane= GetCentreCube(Rotm_face, np.array([x, y, z]), self.MarkerSize, self.distance2markers,None)
                Rotm_Cube = GetOrientationCube(marker_id, RotPlane)
                Rvec_cube, _  = cv2.Rodrigues(Rotm_Cube)
                Rotm_face_prev = Rotm_face
                CentreCube_prev = CentreCube


                Data.append({ 
                    "Corners" : corners[frame_number],
                    "Marker ID" : marker_id,
                    "CentreCube": CentreCube,
                    "CentreCube_alt" : CentreCube_alt,
                    "Rvec_cube": Rvec_cube,
                    "Rvec_cube_alt": np.array([0, 0, 0]),
                    "time": timestamps[frame_number],
                    "frame_number": frame_number,
                    "frame": frame,
                    "Rotframe_prev": np.eye(3),
                    })
                

            else : 
            
                _, rvecs, tvecs, _ = cv2.solvePnPGeneric(self.object_points, image_points, self.CameraMatrix, self.DistCoeffs,flags=cv2.SOLVEPNP_IPPE_SQUARE)

                Rotm_face , Rot_alt ,skip, tvec = GetRightorientation(marker_id, rvecs, tvecs, Rotm_face_prev)
                            
                if self.Pointcloud_avaible :
                        if frame_number >  len(self.depthmap_values)-1:
                            continue

                        depth = self.depthmap_values[frame_number]
                        if depth is None :
                            frame_number += 1
                            continue

                        x,y,z = depth
                        tvec = [x ,y , z]  
                           
                else :
                     x, y, z = tvec.flatten() 
            
                    
                    
               
                #! Orientation:
                # Betterfit:
                RotPlane = GetFacePoints(Rotm_face,np.array([x,y,z]),self.MarkerSize)
                Rotm_Cube = GetOrientationCube(marker_id, RotPlane)
                Rvec_cube, _  = cv2.Rodrigues(Rotm_Cube)
                Rotm_face_prev = Rotm_face

                # for other fit:
                RotPlane_alt = GetFacePoints(Rot_alt,np.array([x,y,z]),self.MarkerSize)
                Rotm_Cube_alt = GetOrientationCube(marker_id, RotPlane_alt)
                Rvec_cube_alt, _  = cv2.Rodrigues(Rotm_Cube_alt)
                

                #! Position:
                CentreCube,CentreCube_alt,RotPlane= GetCentreCube(Rotm_face, np.array([x, y, z]), self.MarkerSize, self.distance2markers,CentreCube_prev)
                CentreCube_prev = CentreCube
                Data.append({
                    "Corners" : image_points,
                    "Marker ID" : marker_id,
                    "CentreCube": CentreCube,
                    "CentreCube_alt" : CentreCube_alt,
                    "Rvec_cube_alt": Rvec_cube_alt,
                    "Rvec_cube": Rvec_cube,
                    "time": timestamps[frame_number],
                    "frame": frame,
                    "frame_number": frame_number,
                    "Rotframe_prev": Rotm_face,
                })
                
            
            frame_number += 1
                

        cap.release()
        self.Data = Data
        return

    def Define_Begin_End(self,idx_start,idx_end):

        remove_frames = list(range(0, idx_start)) + list(range(idx_end, len(self.Data)))
        remove_frames.sort(reverse=True)  # Sort in reverse order to avoid index issues when removing

        for i in remove_frames:
            self.Data.pop(i)
            self.skips.append(i)



    def ReadConfig(self):
        with open(f'./Data_processing/Recordings/{self.name}/configs.txt', 'r') as file:
            lines = file.readlines()

            # Line 0: Pointcloud availability
            depth_line = lines[0]
            if "Pointcloud_avaible" in depth_line:
                self.depthmap = True
            else:
                self.depthmap = False

            # Line 1: Camera quaternion
            second_line = lines[1]
            start = second_line.find('[')
            end = second_line.find(']')
            if start != -1 and end != -1:
                quaternion_str = second_line[start+1:end]
                quaternion = [float(x.strip()) for x in quaternion_str.split(',')]
                self.Cam2world = R.from_quat(quaternion).as_matrix()
            else:
                self.Cam2world = None  # or raise an exception



    

    def Recalculatecpositions_and_orientations(self,idx):
        frame_number =idx

        New_data = self.Data[0:idx-1]


        Rotm_face_prev = self.Data[idx-1]["Rotframe_prev"]
        CentreCube_prev = self.Data[idx-1]["CentreCube"]
        for i in range(idx-1,len(self.Data)):
            image_points = self.Data[i]["Corners"]

            marker_id = self.Data[i]["Marker ID"]


            _, rvecs, tvecs, _ = cv2.solvePnPGeneric(self.object_points, image_points, self.CameraMatrix, self.DistCoeffs,flags=cv2.SOLVEPNP_IPPE_SQUARE)

            Rotm_face , Rot_alt ,skip, tvec = GetRightorientation(marker_id, rvecs, tvecs, Rotm_face_prev)
            if skip : 
                continue

            x, y, z = tvec.flatten() 

            #! Orientation:
            # Betterfit:
            RotPlane = GetFacePoints(Rotm_face,np.array([x,y,z]),self.MarkerSize)
            Rotm_Cube = GetOrientationCube(marker_id, RotPlane)
            Rvec_cube, _  = cv2.Rodrigues(Rotm_Cube)
            

            # for other fit:
            RotPlane_alt = GetFacePoints(Rot_alt,np.array([x,y,z]),self.MarkerSize)
            Rotm_Cube_alt = GetOrientationCube(marker_id, RotPlane_alt)
            Rvec_cube_alt, _  = cv2.Rodrigues(Rotm_Cube_alt)
            

            #! Position:
            CentreCube,CentreCube_alt,RotPlane= GetCentreCube(Rotm_face, np.array([x, y, z]), self.MarkerSize, self.distance2markers,CentreCube_prev)
            CentreCube_prev = CentreCube
            New_data.append({
                "Corners" : image_points,
                "Marker ID" : marker_id,
                "CentreCube": CentreCube,
                "CentreCube_alt" : CentreCube_alt,
                "Rvec_cube_alt": Rvec_cube_alt,
                "Rvec_cube": Rvec_cube,
                "time": self.Data[i]["time"],
                "frame": self.Data[i]["frame"],
                "frame_number": frame_number,
                "Rotframe_prev": Rotm_face_prev,
            })

            Rotm_face_prev = Rotm_face
            frame_number +=1 
        return New_data




    def plot_quats(self):

        rvecs = [item["Rvec_cube"] for item in self.Data]
        rvecs = np.squeeze(rvecs)
        frame_numbers = range(len(self.Data))
        # Unpack current quaternions
        current_quats = R.from_rotvec(rvecs).as_quat()
        qx, qy, qz, qw = zip(*current_quats)
        

        fig, axs = plt.subplots(4, 1, figsize=(14, 10), sharex=True)

        # Left column: quaternion components
        quat_components = {'qx': qx, 'qy': qy, 'qz': qz, 'qw': qw}
        for i, (name, values) in enumerate(quat_components.items()):
            axs[i].plot(frame_numbers, values, label=name, color='blue')
            axs[i].set_ylabel(name)
            axs[i].legend()
            axs[i].grid(True)

        axs[-1].set_xlabel('Frame')
        plt.tight_layout(rect=[0, 0, 1, 0.96])
        plt.show()


    def plot_positions(self):
        frame_numbers = range(len(self.Data))
        positions = [item["CentreCube"] for item in self.Data]
        positions = np.squeeze(positions)
        x, y, z = positions[:, 0], positions[:, 1], positions[:, 2]

        fig, axs = plt.subplots(3, 1, figsize=(14, 8), sharex=True)

        position_components = {'x': x, 'y': y, 'z': z}
        for i, (name, values) in enumerate(position_components.items()):
            axs[i].plot(frame_numbers, values, label=name, color='green')
            axs[i].set_ylabel(f'Position {name}')
            axs[i].legend()
            axs[i].grid(True)

        axs[-1].set_xlabel('Frame')

        plt.tight_layout(rect=[0, 0, 1, 0.96])
        plt.show()


    def Savedata(self):
        if self.exist :
            answer = input('There is allready a safe, do you want to override? [[Y/n]')
            if answer.lower() == 'Y' or answer.upper() == 'Y' :
                self.Saver()
            else :
                print('Aborting saving process')
        else : 
            self.Saver()


    def Saver(self) : 
        
        os.makedirs(f"./Data_processing/savings/{self.name}", exist_ok=True)
        wb = Workbook()
        ws = wb.active

            # Write header
        ws.append(["time", "frame_number", "Marker ID", "CentreCube", "CentreCube_alt", "Rvec_cube", "Rvec_cube_alt", "Rotframe_prev", "Corners"])

        # Write data


        for det in self.Data:
            frame_number = det.get("frame_number")
            timestamp = det.get("time")
            aruco_id = det.get("Marker ID")
            corners = det.get("Corners")
            corners_str = json.dumps(corners.tolist() if hasattr(corners, 'tolist') else corners)

            centre_cube = det.get("CentreCube")
            centre_cube_str = json.dumps(centre_cube.tolist() if hasattr(centre_cube, 'tolist') else centre_cube)

            centre_cube_alt = det.get("CentreCube_alt")
            centre_cube_alt_str = json.dumps(centre_cube_alt.tolist() if hasattr(centre_cube_alt, 'tolist') else centre_cube_alt)

            rvec = det.get("Rvec_cube")
            rvec_str = json.dumps(rvec.tolist() if hasattr(rvec, 'tolist') else rvec)

            rvec_alt = det.get("Rvec_cube_alt")
            rvec_alt_str = json.dumps(rvec_alt.tolist() if hasattr(rvec_alt, 'tolist') else rvec_alt)

            rot_prev = det.get("Rotframe_prev")
            rot_prev_str = json.dumps(rot_prev.tolist() if hasattr(rot_prev, 'tolist') else rot_prev)



            ws.append([
                timestamp,
                frame_number,
                aruco_id,
                centre_cube_str,
                centre_cube_alt_str,
                rvec_str,
                rvec_alt_str,
                rot_prev_str,
                corners_str,
            ])

        # Save the workbook
        wb.save(f"./Data_processing/savings/{self.name}/Data_processed.xlsx")
        print(f"Data saved to ./Data_processing/savings/{self.name}/Data_processed.xlsx")


    def LoadData(self) : 
        cap = cv2.VideoCapture(f'./Data_processing/Recordings/{self.name}/video.mp4')
        frames = []

        while cap.isOpened() :
            ret, frame = cap.read()
            if not ret:
                break
            frames.append(frame)

        if len(self.skips) != 0 :
            for i in self.skips :
                frames.pop(i)


        wb = load_workbook(f"./Data_processing/savings/{self.name}/Data_processed.xlsx")
        ws = wb.active

        # Skip the header
        rows = list(ws.iter_rows(min_row=2, values_only=True)) 


        data = []

        for i, row in enumerate(rows) :
            timestamp, frame_number, aruco_id, centre_cube_str, centre_cube_alt_str, rvec_str, rvec_alt_str, rot_prev_str, corners_str = row
            
            # Convert JSON strings back to original Python structures
            centre_cube = json.loads(centre_cube_str) if centre_cube_str else None
            centre_cube = np.array(centre_cube, dtype=np.float32).reshape((3, 1))
            centre_cube_alt = json.loads(centre_cube_alt_str) if centre_cube_alt_str else None
            centre_cube_alt = np.array(centre_cube_alt, dtype=np.float32).reshape((3, 1))
            rvec = json.loads(rvec_str) if rvec_str else None
            rvec = np.array(rvec, dtype=np.float32).reshape((3, 1))
            rvec_alt = json.loads(rvec_alt_str) if rvec_alt_str else None
            rvec_alt = np.array(rvec_alt, dtype=np.float32).reshape((3, 1))
            rot_prev = json.loads(rot_prev_str) if rot_prev_str else None
            rot_prev = np.array(rot_prev, dtype=np.float32).reshape((3, 3))
            corners = json.loads(corners_str) if corners_str else None
            corners = np.array(corners, dtype=np.float32).reshape((4, 1, 2))


            # Append reconstructed dict
            data.append({
                "frame" : frames[i],
                "time": timestamp,
                "frame_number": frame_number,
                "Marker ID": aruco_id,
                "CentreCube": centre_cube,
                "CentreCube_alt": centre_cube_alt,
                "Rvec_cube": rvec,
                "Rvec_cube_alt": rvec_alt,
                "Rotframe_prev": rot_prev,
                "Corners": corners,
            })

        self.Data = data

    def GetDepthmapdepthdata(self, pixel_coords, original_frames):

        topic_types, reader, type_map = Makereader(self.bagpath)
        depthmap_values = []
        removedfram = 0
        pixel_index = 0
        iteration= -1
        total_needed = len(pixel_coords)

        while reader.has_next() and pixel_index < total_needed:
            topic, data, t = reader.read_next()
            

            
            
            if topic != self.topic_pointcloud:
                continue

            iteration +=1 
            if iteration not in original_frames:
                continue

            msg = deserialize_msg(topic, data, type_map)
            width = msg.width
            height = msg.height

            point_step = msg.point_step

            x_offset = next(f.offset for f in msg.fields if f.name == 'x')
            y_offset = next(f.offset for f in msg.fields if f.name == 'y')
            z_offset = next(f.offset for f in msg.fields if f.name == 'z')

            xu, yv = pixel_coords[pixel_index]
            pixel_index += 1
            u,v = int(np.round(xu*width )), int(np.round(yv*height))
            if not (0 <= u < width and 0 <= v < height):
                print(f"[{pixel_index}] Pixel ({u}, {v}) out of bounds.")
                print(width,height)
                depthmap_values.append(None)
                continue

            index = v * width + u
            offset = index * point_step

            try:
                x = struct.unpack_from('f', msg.data, offset + x_offset)[0]
                y = struct.unpack_from('f', msg.data, offset + y_offset)[0]
                z = struct.unpack_from('f', msg.data, offset + z_offset)[0]
            except struct.error:
                print(f"[{pixel_index}] Corrupt point at ({u}, {v})")
                depthmap_values.append(None)
                continue

            if any(np.isnan([x, y, z])):
                removedfram +=1
                depthmap_values.append(None)
                continue

            depthmap_values.append((x, y, z))
        

        self.depthmap_values = depthmap_values
        print(removedfram)


def distance_to_square_edge(x, y, xmin=0, xmax=0.5, ymin=0, ymax=0.5):
    """
    Compute distance from point (x, y) to the boundary of a square
    from (xmin, ymin) to (xmax, ymax), even if the point lies inside.
    """
    if xmin <= x <= xmax and ymin <= y <= ymax:
        return float(min(x - xmin, xmax - x, y - ymin, ymax - y))
    else:
        dx = float(max(xmin - x, 0, x - xmax))
        dy = float(max(ymin - y, 0, y - ymax))
        return float(np.hypot(dx, dy))  # Ensure scalar float
